from pathlib import Path

from app.converters.base import ConvertResult, rows_to_markdown_table


class XlsxConverter:
    engine = "openpyxl"

    def convert(self, input_path: Path, output_dir: Path) -> ConvertResult:
        from openpyxl import load_workbook

        workbook = load_workbook(input_path, data_only=True, read_only=True)
        parts: list[str] = []
        sheet_metadata: list[dict[str, int | str]] = []

        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            table = rows_to_markdown_table(rows)
            if table:
                parts.append(f"## {sheet.title}\n\n{table}")
            sheet_metadata.append(
                {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }
            )

        workbook.close()

        return ConvertResult(
            markdown="\n\n".join(parts),
            metadata={
                "engine": self.engine,
                "source_format": "xlsx",
                "sheets": sheet_metadata,
            },
        )
